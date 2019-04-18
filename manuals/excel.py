#! python3
#vu6.py - Download Manual views

import glob
import getpass
import os
import progressbar
import sys
import time
from tabulate import tabulate
from pywinauto.timings import wait_until
from pywinauto import findwindows
from datetime import datetime  as dt
from pywinauto.application import Application
from openpyxl import load_workbook
from pathlib import Path

#VARIABLES
ACRONYM = os.environ.get('USERNAME')
FOLDERPATH = os.getcwd()
# EXCELPATH = r'.\manuals.xlsx'
EXCELPATH = sorted(Path('.').glob('*.xlsx'))[0].name

def generate_from_excel(path):
    """generate list modules/part-numbers """
    manual_file_not_in_folder= os.path.exists("manuals.xlsx") is not True
    if manual_file_not_in_folder:
        raise FileNotFoundError
    else:
        wb = load_workbook(path)
        ws = wb['Sheet1']
        maximum_rows = ws.max_row
        column_modules = [ws.cell(row=i+1, column=1).value for i in range(maximum_rows)]
        column_partnumbers = [ws.cell(row=j+1, column=2).value for j in range(maximum_rows)]
        column_modules = list(filter(None, column_modules))
        column_partnumbers = list(filter(None, column_partnumbers))
        return dict(zip(column_modules, column_partnumbers))

def display_board(excel_data):
    """generate a formatted display of content in excel file """
    is_data_dictionnary = isinstance(excel_data, dict) is not True

    if is_data_dictionnary:
        raise TypeError("the value provided to method:display_board should be a dictionnary")
    else:
        modules = list(excel_data.keys())
        partnumbers = list(excel_data.values())
        print(tabulate({"Module": modules, "Part number": partnumbers }, headers="keys"))

def export_view(mod, pt_num, path_to_views, username=ACRONYM, password=ACRONYM):
    cmd_display(mod, 'downloading...')
    app = Application().start(cmd_line=u'"C:\\Program Files (x86)\\Solid Edge TC Manual View\\Solid Edge TC Manual View.exe"')
    app.LoginForm.wait('ready', timeout=30)
    app.LoginForm.Edit5.set_edit_text(username)
    app.LoginForm.Edit4.set_edit_text(password)
    app.LoginForm.Login.click_input()
    time.sleep(5)
    app.PrincipalForm.wait('ready')
    app.PrincipalForm.Edit2.set_edit_text(pt_num)
    app.PrincipalForm.Search.click_input()
    time.sleep(3)
    if findwindows.find_windows(title='Search error'):
        print('Search Error: PT number invalide ')
        window = app.Dialog
        window.wait('ready')
        window.close()
        app.PrincipalForm.type_keys('%{F4}')
    else:
        app.PrincipalForm[u'4'].type_keys(path_to_views) #loop through the  modules files
        app.PrincipalForm[u'3'].type_keys("{DOWN}")
        app.PrincipalForm.Export.click_input()
    app.wait_cpu_usage_lower(threshold=5) # wait until CPU usage is lower than 5%
    def detecting_pdf(): return any(glob.glob(os.path.join(path_to_views, r'*\*.pdf')))
    wait_until(1000, 5.00, detecting_pdf, True)
    cmd_display(mod, 'exported')
    #close the popup windows & app
    window = app.Dialog
    button = window.OK
    button.click_input()
    #close principal form
    app.PrincipalForm.type_keys('%{F4}') # close an active window with Alt+F4
    time.sleep(3)

def cmd_display(module_number, task):
    time_now = dt.now().strftime('%H:%M:%S')
    print(f"\n[\t{module_number}\t]:\t{time_now}\t {task}")

def proceed_yes_or_no():
    "ask user to resume the program"
    answer = input("Proceed ([y]/n) ?:  ")
    if answer in ['yes','y','YES','Y']:
        pass
    else:
        sys.exit("Process has stopped.")

def introduction():
    """Header of the app"""
    today = dt.now().strftime('%H:%M:%S')
    intro = f"""
    Date: {today}
    user: {ACRONYM}
    """
    print(intro)

def password():
    """ request the user password. """
    return getpass.getpass("Enter password :")

def decorator(wrapped_function):
    def _wrapper(*args, **kwargs):
        # do something before the function call
        result = wrapped_function(*args, **kwargs)
        # do something after the function call
        return result
    return _wrapper

def main(excel_path, downloads_path):
    introduction()
    password_user = password()
    tableau = generate_from_excel(excel_path)
    display_board(tableau)
    proceed_yes_or_no()
    for module,number in tableau.items():
        export_view(str(module), str(number), os.path.join(downloads_path, module), password=password_user)  # three variable mandatory

if __name__ == '__main__':
    main(EXCELPATH, FOLDERPATH)
